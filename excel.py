import os
from openpyxl import load_workbook
import win32com.client

workbook = load_workbook('ExoTemplate.xlsx')

feuille = workbook["data"]
feuille.cell(row=1,column=1).value="author"
feuille.cell(row=1,column=2).value="mick"

workbook.save('nouveau_fichier.xlsx')
def exporter_excel_en_pdf(chemin_fichier_excel, chemin_fichier_pdf):
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    wb = excel.Workbooks.Open(chemin_fichier_excel)
    wb.ExportAsFixedFormat(0, chemin_fichier_pdf)

    wb.Close()
    excel.Quit()
FOLDER =r'C:\Users\EPSI\projects\monprojettest'
chemin_fichier_pdf = os.path.join (FOLDER,'chemin_vers_le_fichier.pdf')
chemin_fichier_xls = os.path.join (FOLDER,'nouveau_fichier.xlsx')
exporter_excel_en_pdf(chemin_fichier_xls, chemin_fichier_pdf)